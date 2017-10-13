# -*- coding: utf-8 -*-
#
# Create your app on 'https://ssl.reddit.com/prefs/apps'
# and get the client id and client secret. Add your app name
# and reddit username to user agent:
# user_agent='<app-name>/0.1 by <username>'
# ***(/0.1 is a version of your app, it can be any number)***
# Fill in your username and password from your reddit profile.

import re
import time
import praw
import openpyxl
import datetime
import collections


reddit = praw.Reddit(client_id='',
                     client_secret='',
                     user_agent='',
                     username='',
                     password='')

wb = openpyxl.load_workbook('CoinMarketCap (api).xlsx')  # loading your excel file
ws = wb['Sheet2']
ws['A1'] = None  # deleting first cell with value "Column1.id"

day = 86400  # (24h = 86400 UNIX)
now_time = time.time()  # now time in UNIX

RESULT_DICT = collections.OrderedDict()  # key:value result for each keyword
for (i,) in ws.values:
    if i is not None:
        RESULT_DICT[i] = 0  # creating a dict keyword:0
    else:
        continue


# == Merging subreddits to reduce amount of requests == #


def subreddit():
    global subreddits
    subreddits_list = ['altcoin', 'ethtrader', 'cryptocurrency']  # www.reddit,com/r/<subreddit>
    sub = ""
    for i in subreddits_list:  # adding + between subreddits
        sub += i + "+"
    sub = sub[:-1]  # removing last "+" in string
    subreddits = reddit.subreddit(sub)  # www.reddit.com/r/<subreddit>+<subreddit>+....
    return subreddits


# ========== Search engine ========== #


def search_engine(text, title):
    for (i,) in ws.values:
        if i is not None:  # check for None value in cell
            keyword = str(i).replace('-', '.')
            pattern = re.compile(keyword, re.IGNORECASE)
            result_text = re.findall(pattern, text)  # searching for the first match in text
            result_title = re.findall(pattern, title)
            if result_text is not None:
                RESULT_DICT[str(i)] += len(result_text)  # + number of matches to value if keyword will be found
            if result_title is not None:
                RESULT_DICT[str(i)] += len(result_title)
        else:
            pass


# === Search for submissions in given subreddits === #


def submission_search():
    for submission in subreddits.submissions(start=(now_time-day)):  # Getting all subreddits for the last 24h
        title = submission.title.encode('utf-8')  # submission title
        text = submission.selftext.encode('utf-8')  # submission text
        created = time.ctime(submission.created_utc)  # created time
        print 20 * '='
        print created
        print 20 * "-"
        print title
        print text
        search_engine(text, title)


# === Search for comments in given subreddits === #


def comments_search():
    for submission in subreddits.submissions(start=(now_time-day*3)):  # Getting all subreddits for the last 72h
        submission.comments.replace_more(limit=0)  # searching for all comments and replies
        for comment in submission.comments.list():
            if comment.created_utc >= (period-day):
                search_engine(comment.body, title='')
            text = comment.body.encode('utf-8')  # comment text
            created = time.ctime(comment.created_utc)  # created time
            print 20 * '-'
            print created
            print text


# ========== Saving the result to excel ========== #


def save_results():
    now_date = datetime.datetime.now().strftime('%Y-%m-%d')  # now date
    workbook = openpyxl.Workbook()  # creating new file
    sheet = workbook.active
    items = RESULT_DICT.items()  # getting all key:values
    for (i, k) in items:
        result = (i, k)
        sheet.append(result)  # writing results
    workbook.save('CoinMarketCap(%s).xlsx' % now_date)  # save file as "CoinMarketCap(<now date>).xlsx"
    workbook.close()


if __name__ == "__main__":
    subreddit()
    submission_search()
    comments_search()
    print RESULT_DICT
    save_results()

















































































